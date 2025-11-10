# invapp/views/party_views.py

from django.shortcuts import render, get_object_or_404, redirect
from brokerapp.forms import PartyForm, BrokerForm, ItemForm
from brokerapp.models import HeadParty, Broker, HeadItem ,SaleMaster, SaleDetails ,PurchaseMaster, PurchaseDetails, DailyPage, JamaEntry, NaameEntry
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Max
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse
import json
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from django.utils import timezone
from django.db.models import Sum, Prefetch, F, FloatField, ExpressionWrapper
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.db.models import ProtectedError
from io import BytesIO
from fpdf import FPDF
from django.views.generic import TemplateView
from .forms import AllPartyBalanceForm
from django.urls import reverse
from urllib.parse import quote
import io

class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "Daily Report", ln=1, align="C")
        self.ln(5)

    def table_side(self, title, entries, total, x_pos, max_rows):
        """Side table banane ke liye"""
        start_y = self.get_y()
        self.set_xy(x_pos, start_y)
        self.set_font("Arial", "B", 12)
        self.cell(90, 10, title, ln=1, align="L")

        # header
        self.set_x(x_pos)
        self.set_font("Arial", "B", 10)
        self.cell(20, 8, "No", 1, 0, "C")
        self.cell(40, 8, "Party", 1, 0, "C")
        self.cell(30, 8, "Amount", 1, 1, "C")

        # rows
        self.set_font("Arial", "", 10)
        for e in entries:
            self.set_x(x_pos)
            self.cell(20, 8, str(e.entry_no), 1, 0, "C")
            self.cell(40, 8, e.party.partyname, 1, 0)
            self.cell(30, 8, f"{e.amount:.2f}", 1, 1, "R")

        # padding blank rows
        extra_rows = max_rows - len(entries)
        for _ in range(extra_rows):
            self.set_x(x_pos)
            self.cell(20, 8, "", 1, 0, "C")
            self.cell(40, 8, "", 1, 0)
            self.cell(30, 8, "", 1, 1, "R")

        # total row
        self.set_x(x_pos)
        self.set_font("Arial", "B", 10)
        self.cell(60, 8, "Total", 1, 0, "R")
        self.cell(30, 8, f"{total:.2f}", 1, 1, "R")



# -----------------------
# Helpers
# -----------------------
def to_decimal(val, default=Decimal('0')):
    """Safe conversion to Decimal."""
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError, ValueError):
        return default




# -----------------------
# Views
# -----------------------

def sale_form(request, invno=None):
    """
    Render sale form. If invno provided, load sale and its details (scoped to current org).
    """
    sale = None
    sale_items_json = "[]"
    today_date = date.today().strftime("%Y-%m-%d")

    if invno:
        # sale must belong to current org
        sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)
        details = SaleDetails.objects.filter(salemaster=sale)
        items_data = []
        for d in details:
            items_data.append({
                "item_id": d.item.pk,
                "item_name": d.item.item_name,
                "bora": float(d.bora),
                "bn": float(d.bn),
                "bnwt": float(d.bnwt),
                "bo": float(d.bo),
                "bowt": float(d.bowt),
                "tbwt": float(getattr(d, "tbwt", 0)),
                "totalbora": float(d.bn*d.bnwt + d.bo*d.bowt),
                "qty": float(d.qty),
                "rate": float(d.rate),
                "amt": float(d.amount),
                "partywt": float(d.partywt),
                "millwt": float(d.millwt),
                "diffwt": float(d.diffwt),
                "lotno": d.lotno or "",
            })
        sale_items_json = json.dumps(items_data)

    # next invoice number — per ORG
    next_invno = SaleMaster.objects.filter(org=request.current_org).aggregate(Max("invno"))['invno__max']
    next_invno = (next_invno + 1) if next_invno else 1

    context = {
        "sale": sale,
        "sale_items_json": sale_items_json,
        "next_invno": next_invno,
        "today_date": today_date,
        # only current org choices
        "items": HeadItem.objects.filter(org=request.current_org).order_by('item_name'),
        "parties": HeadParty.objects.filter(org=request.current_org).order_by('partyname'),
        "brokers": Broker.objects.filter(org=request.current_org).order_by('brokername'),
    }
    return render(request, "brokerapp/sale.html", context)

# ===================== SAVE SALE =====================

@transaction.atomic
def save_sale(request):
    """
    Save a new SaleMaster and its SaleDetails — scoped to current org.
    """
    if request.method != "POST":
        return redirect('sale_form_new')

    try:
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else date.today()
        awakno = request.POST.get("awakno", "").strip()
        extra = request.POST.get("extra", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()

        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("sale_form_new")

        total_amt = Decimal('0')
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal('100')).quantize(Decimal('0.01'))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal('100')).quantize(Decimal('0.01'))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))
        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal('0.01'))
        netamt = (total - advance).quantize(Decimal('0.01'))

        # Resolve FKs inside same org
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        # Create SaleMaster with org + created_by
        sale = SaleMaster.objects.create(
            org=request.current_org,
            created_by=request.user,
            invdate=invdate,
            awakno=awakno,
            party=party,
            broker=broker,
            vehicleno=vehicleno,
            extra=extra,
            totalamt=total_amt.quantize(Decimal('0.01')),
            batavpercent=batavpercent,
            batavamt=batavamt,
            dr=dr,
            dramt=dramt,
            qi=qi,
            other=other,
            total=total,
            advance=advance,
            netamt=netamt,
            remark=request.POST.get("remark", "").strip(),
        )

        # Create SaleDetails (items limited to same org)
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)
            SaleDetails.objects.create(
                salemaster=sale,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        messages.success(request, "Sale entry saved successfully!")
        return redirect("saledata")

    except Exception as e:
        messages.error(request, f"Error saving sale: {e}")
        return redirect("sale_form_new")


@transaction.atomic
def update_sale(request, invno):
    """
    Update existing SaleMaster identified by invno (scoped to current org).
    """
    sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)

    if request.method != "POST":
        return redirect('sale_form_update', invno=invno)

    try:
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else sale.invdate
        awakno = request.POST.get("awakno", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()
        extra = request.POST.get("extra", "").strip()

        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("sale_form_update", invno=invno)

        total_amt = Decimal('0')
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal('100')).quantize(Decimal('0.01'))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal('100')).quantize(Decimal('0.01'))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))
        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal('0.01'))
        netamt = (total - advance).quantize(Decimal('0.01'))

        # Resolve FKs inside same org
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        # Update header
        sale.invdate = invdate
        sale.awakno = awakno
        sale.party = party
        sale.broker = broker
        sale.vehicleno = vehicleno
        sale.extra = extra
        sale.totalamt = total_amt.quantize(Decimal('0.01'))
        sale.batavpercent = batavpercent
        sale.batavamt = batavamt
        sale.dr = dr
        sale.dramt = dramt
        sale.qi = qi
        sale.other = other
        sale.advance = advance
        sale.total = total
        sale.netamt = netamt
        sale.remark = request.POST.get("remark", "").strip()
        sale.save()

        # Replace details
        SaleDetails.objects.filter(salemaster=sale).delete()
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)
            SaleDetails.objects.create(
                salemaster=sale,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        messages.success(request, "Sale entry updated successfully!")
        return redirect("saledata")

    except Exception as e:
        messages.error(request, f"Error updating sale: {e}")
        return redirect("sale_form_update", invno=invno)


def sale_data_view(request):
    """List of sales (scoped to current org)."""
    sales = SaleMaster.objects.filter(org=request.current_org).order_by("-invno")
    return render(request, "brokerapp/saledata.html", {
        "sales": sales,
        "today_date": date.today(),
    })




def delete_sale(request, invno):
    sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)
    sale.delete()
    messages.success(request, "Sale entry deleted successfully!")
    return redirect("saledata")




def sale_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    broker_id = request.GET.get("broker")
    report_type = request.GET.get("report_type", "date")

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset (ORG SCOPED) + prefetch details (with item) for the template
    sales = (
        SaleMaster.objects
        .filter(org=request.current_org)
        .select_related("broker")
        .prefetch_related(Prefetch("details", queryset=SaleDetails.objects.select_related("item")))
    )

    if start_date:
        sales = sales.filter(invdate__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(invdate__lte=parse_date(end_date))

    # Broker filter: allow pk or name, but scoped to org
    if broker_id and broker_id != "all":
        # try as primary key
        if sales.filter(broker__pk=broker_id).exists():
            sales = sales.filter(broker__pk=broker_id)
        else:
            sales = sales.filter(broker__brokername=broker_id)

    sales = sales.order_by("invdate")

    report_data = []

    # --- Grouping & Aggregates ---
    if report_type == "date":
        grouped = sales.values("invdate").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate")

        for g in grouped:
            group_sales = sales.filter(invdate=g["invdate"])
            # TBWt sum for this group (sum over details)
            tbwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
            g["total_tbwt"] = tbwt_sum
            report_data.append({
                "group": g["invdate"],
                "items": group_sales,
                "totals": g
            })

    elif report_type == "broker":
        grouped = sales.values("invdate", "broker__brokername").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "broker__brokername")

        for g in grouped:
            group_sales = sales.filter(
                invdate=g["invdate"],
                broker__brokername=g["broker__brokername"]
            )
            tbwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
            g["total_tbwt"] = tbwt_sum
            report_data.append({
                "group": f"{g['invdate']} - {g['broker__brokername'] or 'No Broker'}",
                "items": group_sales,
                "totals": g
            })

    # Overall Totals (header-level) + TBWt across all details in the filtered set
    overall_totals = sales.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_total=Sum("total"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
    overall_totals["total_tbwt"] = overall_tbwt

    # Dropdowns also ORG SCOPED
    brokers = Broker.objects.filter(org=request.current_org).order_by("brokername")

    context = {
        "report_data": report_data,
        "overall_totals": overall_totals,
        "start_date": start_date,
        "end_date": end_date,
        "brokers": brokers,
        "selected_broker": broker_id if broker_id != "all" else None,
        "report_type": report_type,
    }
    return render(request, "brokerapp/sale_report.html", context)

# ===================== PDF (FPDF) =====================
def sale_report_pdf(request):
    """
    Generate Sale Report PDF (FPDF) using current filters.
    Includes per-invoice detail rows with TBWt.
    Numbers are right-aligned with thousand separators.
    """
    # --- build same filtered queryset as HTML report ---
    start_date = request.GET.get("start_date") or date.today().strftime("%Y-%m-%d")
    end_date = request.GET.get("end_date") or date.today().strftime("%Y-%m-%d")
    broker_id = request.GET.get("broker")
    report_type = request.GET.get("report_type", "date")

    sales = (
        SaleMaster.objects
        .filter(org=request.current_org)
        .select_related("broker")
        .prefetch_related(Prefetch("details", queryset=SaleDetails.objects.select_related("item")))
    )
    if start_date:
        sales = sales.filter(invdate__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(invdate__lte=parse_date(end_date))
    if broker_id and broker_id != "all":
        if sales.filter(broker__pk=broker_id).exists():
            sales = sales.filter(broker__pk=broker_id)
        else:
            sales = sales.filter(broker__brokername=broker_id)
    sales = sales.order_by("invdate", "invno")

    # group-key helpers (just for headings)
    if report_type == "date":
        def group_key(s): return (s.invdate,)
    else:
        def group_key(s): return (s.invdate, s.broker.brokername if s.broker else "")

    # --- FPDF setup ---
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Sale Report", ln=1, align="C")
    pdf.set_font("Helvetica", "", 9)
    hdr = f"From {start_date} To {end_date} | Generated: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}"
    pdf.cell(0, 6, hdr, ln=1, align="C")
    pdf.ln(2)

    # --------- MICRO-POLISH HELPERS ----------
    def fmt2(v):
        """format number with commas & 2 decimals"""
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return "0.00"

    def cellR(w, h, txt, **kw):
        """right-aligned numeric cell"""
        pdf.cell(w, h, txt, align="R", **kw)
    # -----------------------------------------

    # headers
    def draw_invoice_header():
        pdf.set_fill_color(230, 240, 255)
        pdf.set_font("Helvetica", "B", 9)
        cols = [
            ("Inv No", 20), ("Date", 22), ("Broker", 40),
            ("Total", 22), ("Batav", 22), ("DR", 18),
            ("Other", 18), ("Adv", 18), ("Net", 22),
        ]
        for text, w in cols:
            pdf.cell(w, 7, text, border=1, align="C", fill=True)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 9)

    def draw_detail_header():
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("Helvetica", "B", 8)
        cols = [
            ("Item", 44), ("Bora", 16), ("TBWt", 16),
            ("Qty", 16), ("Rate", 16), ("Amount", 22),
            ("PWt", 16), ("MWt", 16), ("DWt", 16), ("Lot", 16),
        ]
        for text, w in cols:
            pdf.cell(w, 6, text, border=1, align="C", fill=True)
        pdf.ln(6)
        pdf.set_font("Helvetica", "", 8)

    current_group = None
    draw_invoice_header()

    for s in sales:
        key = group_key(s)
        if current_group is None or key != current_group:
            # group band
            pdf.set_font("Helvetica", "B", 9)
            if report_type == "date":
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')}"
            else:
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')} - {key[1] or 'No Broker'}"
            pdf.ln(2)
            pdf.set_fill_color(235, 235, 235)
            pdf.cell(0, 6, grp_txt, ln=1, fill=True)
            pdf.set_font("Helvetica", "", 9)
            current_group = key

        # invoice header row: text left, numbers right
        pdf.cell(20, 7, str(s.invno), border=1, align="C")
        pdf.cell(22, 7, s.invdate.strftime("%d-%m-%Y"), border=1, align="C")
        pdf.cell(40, 7, (s.broker.brokername if s.broker else "")[:20], border=1, align="L")
        cellR(22, 7, fmt2(s.totalamt), border=1)
        cellR(22, 7, fmt2(s.batavamt), border=1)
        cellR(18, 7, fmt2(s.dramt), border=1)
        cellR(18, 7, fmt2(s.other), border=1)
        cellR(18, 7, fmt2(s.advance), border=1)
        cellR(22, 7, fmt2(s.netamt), border=1)
        pdf.ln(7)

        # details
        draw_detail_header()
        for d in s.details.all():
            pdf.cell(44, 6, (d.item.item_name or "")[:28], border=1, align="L")
            cellR(16, 6, fmt2(d.bora), border=1)
            cellR(16, 6, fmt2(d.tbwt), border=1)
            cellR(16, 6, fmt2(d.qty), border=1)
            cellR(16, 6, fmt2(d.rate), border=1)
            cellR(22, 6, fmt2(d.amount), border=1)
            cellR(16, 6, fmt2(d.partywt), border=1)
            cellR(16, 6, fmt2(d.millwt), border=1)
            cellR(16, 6, fmt2(d.diffwt), border=1)
            pdf.cell(16, 6, (d.lotno or "")[:8], border=1, align="C")
            pdf.ln(6)

    # overall totals
    overall = sales.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(
        total_tbwt=Sum("tbwt")
    )["total_tbwt"] or 0

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Overall Totals", ln=1)
    pdf.set_font("Helvetica", "", 9)
    lines = [
        f"Total Amt: {fmt2(overall['total_totalamt'] or 0)}",
        f"Batav Amt: {fmt2(overall['total_batavamt'] or 0)}",
        f"DR Amt: {fmt2(overall['total_dramt'] or 0)}",
        f"Other: {fmt2(overall['total_other'] or 0)}",
        f"Advance: {fmt2(overall['total_advance'] or 0)}",
        f"Total TBWt: {fmt2(overall_tbwt)}",
        f"Net Amt: {fmt2(overall['total_netamt'] or 0)}",
    ]
    for line in lines:
        pdf.cell(0, 6, line, ln=1)

    # finalize (bytes -> HttpResponse)
    pdf.alias_nb_pages()
    filename = f"sale_report_{start_date}_{end_date}.pdf"
    pdf_bytes = pdf.output(dest="S").encode("latin-1", "replace")
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp



def bardana_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    party_id = request.GET.get("party")
    broker_id = request.GET.get("broker")
    report_type = request.GET.get("report_type", "date")  # date / party / broker

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset (ORG SCOPED via salemaster__org)
    details = (
        SaleDetails.objects
        .select_related('salemaster', 'item', 'salemaster__party', 'salemaster__broker')
        .filter(
            salemaster__org=request.current_org,
            salemaster__invdate__gte=parse_date(start_date),
            salemaster__invdate__lte=parse_date(end_date)
        )
        .order_by('salemaster__invdate')
    )

    # Party / Broker filters (within org)
    if party_id and party_id != "all":
        details = details.filter(salemaster__party__pk=party_id)
    if broker_id and broker_id != "all":
        # allow pk or name
        if details.filter(salemaster__broker__pk=broker_id).exists():
            details = details.filter(salemaster__broker__pk=broker_id)
        else:
            details = details.filter(salemaster__broker__brokername=broker_id)

    # Prepare grouped data
    report_data = []

    if report_type == "date":
        group_values = details.values_list('salemaster__invdate', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__invdate=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g.strftime("%d-%m-%Y"),
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    elif report_type == "party":
        group_values = details.values_list('salemaster__party__partyname', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__party__partyname=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g,
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    elif report_type == "broker":
        group_values = details.values_list('salemaster__broker__brokername', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__broker__brokername=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g or "No Broker",
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    # ORG-scoped dropdown lists
    parties = HeadParty.objects.filter(org=request.current_org).order_by("partyname")
    brokers = Broker.objects.filter(org=request.current_org).order_by("brokername")

    context = {
        "report_data": report_data,
        "parties": parties,
        "brokers": brokers,
        "start_date": start_date,
        "end_date": end_date,
        "selected_party": party_id if party_id != "all" else None,
        "selected_broker": broker_id if broker_id != "all" else None,
        "report_type": report_type,
    }
    return render(request, "brokerapp/bardana_report.html", context)


def purchase_form(request, invno=None):
    """
    Render purchase form. If invno provided, load purchase + details (scoped to current org).
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = None
    purchase_items_json = "[]"
    today_date = date.today().strftime("%Y-%m-%d")

    if invno:
        # must belong to current org
        purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)
        details = PurchaseDetails.objects.filter(purchasemaster=purchase)
        items_data = []
        for d in details:
            items_data.append({
                "item_id": d.item.pk,
                "item_name": d.item.item_name,
                "bora": float(d.bora),
                "bn": float(d.bn),
                "bnwt": float(d.bnwt),
                "bo": float(d.bo),
                "bowt": float(d.bowt),
                "totalbora": float(d.bn*d.bnwt + d.bo*d.bowt),
                "qty": float(d.qty),
                "rate": float(d.rate),
                "amt": float(d.amount),
                "partywt": float(d.partywt),
                "millwt": float(d.millwt),
                "diffwt": float(d.diffwt),
                "lotno": d.lotno or "",
            })
        purchase_items_json = json.dumps(items_data)

    # next invoice number — per ORG
    next_invno = PurchaseMaster.objects.filter(org=request.current_org).aggregate(Max("invno"))['invno__max']
    next_invno = (next_invno + 1) if next_invno else 1

    context = {
        "purchase": purchase,
        "purchase_items_json": purchase_items_json,
        "next_invno": next_invno,
        "today_date": today_date,
        # only current org choices
        "items": HeadItem.objects.filter(org=request.current_org).order_by('item_name'),
        "parties": HeadParty.objects.filter(org=request.current_org).order_by('partyname'),
        "brokers": Broker.objects.filter(org=request.current_org).order_by('brokername'),
    }
    return render(request, "brokerapp/purchase.html", context)



@transaction.atomic
def save_purchase(request):
    """
    Save a new PurchaseMaster and its PurchaseDetails — scoped to current org.
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    if request.method != "POST":
        return redirect('purchase_form_new')

    try:
        # Header
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else date.today()
        awakno = request.POST.get("awakno", "").strip()
        extra = request.POST.get("extra", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()

        # Items
        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("purchase_form_new")

        # Totals
        total_amt = Decimal('0')
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal('100')).quantize(Decimal('0.01'))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal('100')).quantize(Decimal('0.01'))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))
        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal('0.01'))
        netamt = (total - advance).quantize(Decimal('0.01'))

        # Resolve FKs within same org
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        # Create master (bind org + created_by)
        purchase = PurchaseMaster.objects.create(
            org=request.current_org,
            created_by=request.user,
            invdate=invdate,
            awakno=awakno,
            party=party,
            broker=broker,
            vehicleno=vehicleno,
            extra=extra,
            totalamt=total_amt.quantize(Decimal('0.01')),
            batavpercent=batavpercent,
            batavamt=batavamt,
            dr=dr,
            dramt=dramt,
            qi=qi,
            other=other,
            total=total,
            advance=advance,
            netamt=netamt,
            remark=request.POST.get("remark", "").strip(),
        )

        # Create details (items only from same org)
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)
            PurchaseDetails.objects.create(
                purchasemaster=purchase,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        messages.success(request, "Purchase entry saved successfully!")
        return redirect("purchasedata")

    except Exception as e:
        messages.error(request, f"Error saving purchase: {e}")
        return redirect("purchase_form_new")



@transaction.atomic
def update_purchase(request, invno):
    """
    Update existing PurchaseMaster (scoped to current org).
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)

    if request.method != "POST":
        return redirect('purchase_form_update', invno=invno)

    try:
        # Header
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else purchase.invdate
        awakno = request.POST.get("awakno", "").strip()
        extra = request.POST.get("extra", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()

        # Items
        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("purchase_form_update", invno=invno)

        # Totals
        total_amt = Decimal('0')
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal('100')).quantize(Decimal('0.01'))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal('100')).quantize(Decimal('0.01'))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))
        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal('0.01'))
        netamt = (total - advance).quantize(Decimal('0.01'))

        # Resolve FKs in same org
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        # Update master
        purchase.invdate = invdate
        purchase.awakno = awakno
        purchase.extra = extra
        purchase.party = party
        purchase.broker = broker
        purchase.vehicleno = vehicleno
        purchase.totalamt = total_amt.quantize(Decimal('0.01'))
        purchase.batavpercent = batavpercent
        purchase.batavamt = batavamt
        purchase.dr = dr
        purchase.dramt = dramt
        purchase.qi = qi
        purchase.other = other
        purchase.total = total
        purchase.advance = advance
        purchase.netamt = netamt
        purchase.remark = request.POST.get("remark", "").strip()
        purchase.save()

        # Replace details
        PurchaseDetails.objects.filter(purchasemaster=purchase).delete()
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)
            PurchaseDetails.objects.create(
                purchasemaster=purchase,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        messages.success(request, "Purchase entry updated successfully!")
        return redirect("purchasedata")

    except Exception as e:
        messages.error(request, f"Error updating purchase: {e}")
        return redirect("purchase_form_update", invno=invno)


def purchase_data_view(request):
    """List of purchases (scoped to current org)."""
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchases = PurchaseMaster.objects.filter(org=request.current_org).order_by("-invno")
    return render(request, "brokerapp/purchasedata.html", {
        "purchases": purchases,
        "today_date": date.today(),
    })




def delete_purchase(request, invno):
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)
    purchase.delete()
    messages.success(request, "Purchase entry deleted successfully!")
    return redirect("purchasedata")

def purchase_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    broker_id = request.GET.get("broker")
    report_type = request.GET.get("report_type", "date")

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset
    purchases = PurchaseMaster.objects.all()
    if start_date:
        purchases = purchases.filter(invdate__gte=parse_date(start_date))
    if end_date:
        purchases = purchases.filter(invdate__lte=parse_date(end_date))
    if broker_id and broker_id != "all":
        purchases = purchases.filter(broker__brokername=broker_id)

    purchases = purchases.order_by("invdate")

    report_data = []

    # --- Grouping & Aggregates ---
    if report_type == "date":
        grouped = purchases.values("invdate").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate")

        for g in grouped:
            group_purchases = purchases.filter(invdate=g["invdate"])
            report_data.append({
                "group": g["invdate"],
                "items": group_purchases,
                "totals": g
            })

    elif report_type == "broker":
        grouped = purchases.values("invdate", "broker__brokername").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "broker__brokername")

        for g in grouped:
            group_purchases = purchases.filter(
                invdate=g["invdate"],
                broker__brokername=g["broker__brokername"]
            )
            report_data.append({
                "group": f"{g['invdate']} - {g['broker__brokername'] or 'No Broker'}",
                "items": group_purchases,
                "totals": g
            })

    # Overall Totals
    overall_totals = purchases.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_total=Sum("total"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )

    brokers = Broker.objects.all()

    context = {
        "report_data": report_data,
        "overall_totals": overall_totals,
        "start_date": start_date,
        "end_date": end_date,
        "brokers": brokers,
        "selected_broker": broker_id if broker_id != "all" else None,
        "report_type": report_type,
    }

    return render(request, "brokerapp/purchase_report.html", context)



def party_view(request, pk=None):
    # Only fetch inside current org
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(HeadParty, pk=pk, org=request.current_org)

    if request.method == 'POST':
        # ⬇️ current_org पास करें
        form = PartyForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)

            # Ensure party always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Party updated successfully!')
            else:
                messages.success(request, '✅ Party added successfully!')

            # ✅ Redirect logic simplified for Sale return
            next_page = request.GET.get('next')
            if next_page == 'sale':
                # user came from Sale page → go back to Sale after saving
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Party list (existing behavior)
            return redirect(f"{reverse('party')}?created_id={obj.pk}&created_name={quote(obj.partyname)}")

    else:
        # ⬇️ current_org पास करें
        form = PartyForm(instance=instance, current_org=request.current_org)

    # Show only current org parties
    parties = HeadParty.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/party.html', {
        'form': form,
        'parties': parties,
        'editing': pk is not None,
        'editing_id': pk
    })



def party_delete(request, pk):
    """Safely delete a Party — show message if linked to transactions."""
    party = get_object_or_404(HeadParty, pk=pk)
    party_name = party.partyname  # ✅ Save name before deleting
    try:
        party.delete()
        messages.success(request, f"✅ Party '{party_name}' deleted successfully!")
    except ProtectedError:
        messages.error(
            request,
            f"⚠️ Cannot delete '{party_name}' — it is linked to existing Jama or other entries."
        )
    return redirect('party')


def broker_view(request, pk=None):
    # Only fetch inside current org (same as party_view)
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(Broker, pk=pk, org=request.current_org)

    if request.method == 'POST':
        # ⬇️ pass current_org into form
        form = BrokerForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)

            # Ensure broker always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Broker updated successfully!')
            else:
                messages.success(request, '✅ Broker added successfully!')

            # ✅ Redirect logic simplified for Sale return
            next_page = request.GET.get('next')
            if next_page == 'sale':
                # user came from Sale page → go back to Sale after saving
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Broker list (existing behavior)
            return redirect(f"{reverse('broker')}?created_id={obj.pk}&created_name={quote(obj.brokername)}")

    else:
        # ⬇️ pass current_org into form
        form = BrokerForm(instance=instance, current_org=request.current_org)

    # Show only current org brokers
    brokers = Broker.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/broker.html', {
        'form': form,
        'brokers': brokers,
        'editing': pk is not None,
        'editing_id': pk
    })
# Delete Broker
def broker_delete(request, pk):
    """Safely delete a Broker — show warning if linked to transactions."""
    broker = get_object_or_404(Broker, pk=pk)
    broker_name = broker.brokername  # ✅ store name before delete

    try:
        broker.delete()
        messages.success(request, f"✅ Broker '{broker_name}' deleted successfully!")
    except ProtectedError:
        messages.error(
            request,
            f"⚠️ Cannot delete '{broker_name}' — it is linked to existing Jama or Naame entries."
        )

    return redirect('broker')


@login_required
def dashboard(request):
    return render(request, 'brokerapp/dashboard.html')


def item_view(request, pk=None):
    # Only fetch inside current org (same as party_view / broker_view)
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(HeadItem, pk=pk, org=request.current_org)

    if request.method == 'POST':
        action = request.POST.get('action')

        # ---- DELETE action ----
        if action == 'delete':
            # item_pk may be item_name (string PK) or provided via URL pk
            item_pk = request.POST.get('item_pk') or pk
            if not item_pk:
                messages.error(request, '❌ No item selected to delete.')
                return redirect(reverse('item'))

            try:
                obj_to_delete = get_object_or_404(HeadItem, pk=item_pk, org=request.current_org)
                obj_to_delete.delete()
                messages.success(request, '✅ Item deleted successfully!')
            except Exception as e:
                # catch FK/constraint errors or unexpected issues
                messages.error(request, f'❌ Could not delete item: {e}')
            return redirect(reverse('item'))

        # ---- SAVE / CREATE / UPDATE ----
        form = ItemForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)
            # Ensure item always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Item updated successfully!')
            else:
                messages.success(request, '✅ Item added successfully!')

            # Redirect back to Sale if requested
            next_page = request.GET.get('next')
            if next_page == 'sale':
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Item list and pass created id/name for JS
            return redirect(f"{reverse('item')}?created_id={obj.pk}&created_name={quote(obj.item_name)}")
        else:
            messages.error(request, f"❌ Could not save item:\n{form.errors.as_text()}")
    else:
        # pass current_org into form on GET as well
        form = ItemForm(instance=instance, current_org=request.current_org)

    # Show only current org items
    items = HeadItem.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/item.html', {
        'form': form,
        'items': items,
        'editing': pk is not None,
        'editing_id': pk
    })




@require_GET

def daily_page_view(request):
    """
    Show daily page for selected date (via ?date=YYYY-MM-DD) scoped to current org.
    If no date provided, default to today.
    """
    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    # Only current org data
    parties = HeadParty.objects.filter(org=request.current_org).order_by('partyname')
    brokers = Broker.objects.filter(org=request.current_org).order_by('brokername')

    # DailyPage for current org + selected date
    daily_page = DailyPage.objects.filter(
        org=request.current_org,
        date=selected_date
    ).first()

    def serialize_entry(e):
        broker_name = getattr(e.broker, 'brokername', '') if getattr(e, 'broker', None) else ''
        party_name = getattr(e.party, 'partyname', '') if getattr(e, 'party', None) else ''
        return {
            'entry_no': e.entry_no,
            'party_name': party_name,
            'broker_name': broker_name,
            'amount': e.amount,
            'remark': e.remark,
        }

    jama_entries = []
    naame_entries = []

    if daily_page:
        jama_entries = [serialize_entry(j) for j in daily_page.jama_entries.all()]
        naame_entries = [serialize_entry(n) for n in daily_page.naame_entries.all()]

    no_entries = not (jama_entries or naame_entries)

    context = {
        'selected_date': selected_date,
        'parties': parties,
        'brokers': brokers,
        'jama_entries': jama_entries,
        'naame_entries': naame_entries,
        'no_entries': no_entries,
    }
    return render(request, 'brokerapp/daily_page.html', context)

@require_GET

def daily_page_show(request):
    """
    JSON endpoint: ?date=YYYY-MM-DD (optional; if missing -> today)
    Response: { "date": "...", "jama": [...], "naame": [...] }
    """
    d = request.GET.get('date', '').strip()

    # default to today if missing or invalid
    if not d:
        date_obj = timezone.localdate()
    else:
        try:
            date_obj = datetime.strptime(d, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'invalid date format, expected YYYY-MM-DD'}, status=400)

    # ⬇️ scope to current org
    daily_page = DailyPage.objects.filter(org=request.current_org, date=date_obj).first()

    def serialize_entry(entry):
        broker_name = getattr(entry.broker, 'brokername', '') if getattr(entry, 'broker', None) else ''
        party_name = getattr(entry.party, 'partyname', '') if getattr(entry, 'party', None) else ''
        return {
            'entry_no': entry.entry_no,
            'party_name': party_name,
            'broker_name': broker_name,
            'amount': float(entry.amount or 0),
            'remark': entry.remark or '',
            'created_at': entry.created_at.isoformat() if entry.created_at else None,
        }

    jama = []
    naame = []

    if daily_page:
        jama = [serialize_entry(j) for j in daily_page.jama_entries.all()]
        naame = [serialize_entry(n) for n in daily_page.naame_entries.all()]

    if not jama and not naame:
        return JsonResponse({
            'date': date_obj.strftime('%Y-%m-%d'),
            'message': 'No entry on that day',
            'jama': [],
            'naame': [],
        })

    return JsonResponse({
        'date': date_obj.strftime('%Y-%m-%d'),
        'jama': jama,
        'naame': naame,
    })

@require_POST

def daily_page_jama_add(request):
    # expects: date, party (pk), broker (pk), amount, remark (optional)
    date_str = request.POST.get('date')
    party_id = request.POST.get('party')
    broker_id = request.POST.get('broker')
    amount = request.POST.get('amount')
    remark = (request.POST.get('remark') or '').strip()

    if not (date_str and party_id and broker_id and amount):
        return JsonResponse({'error': 'Missing fields'}, status=400)

    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        amt = float(amount)
    except Exception:
        return JsonResponse({'error': 'Invalid input'}, status=400)

    # ⬇️ Party/Broker must belong to current org
    party = get_object_or_404(HeadParty, pk=party_id, org=request.current_org)
    broker = get_object_or_404(Broker, pk=broker_id, org=request.current_org)

    with transaction.atomic():
        # ⬇️ DailyPage is per (org, date)
        daily_page, _ = DailyPage.objects.get_or_create(org=request.current_org, date=date_obj)
        entry = JamaEntry.objects.create(
            daily_page=daily_page,
            party=party,
            broker=broker,
            amount=amt,
            remark=remark
        )

    data = {
        'entry_no': entry.entry_no,
        'party_name': party.partyname,
        'broker_name': broker.brokername,
        'amount': f"{entry.amount:.2f}",
        'remark': entry.remark,
    }
    return JsonResponse({'success': True, 'entry': data})


@require_POST

def daily_page_naame_add(request):
    # expects: date, party (pk), broker (pk), amount, remark (optional)
    date_str = request.POST.get('date')
    party_id = request.POST.get('party')
    broker_id = request.POST.get('broker')
    amount = request.POST.get('amount')
    remark = (request.POST.get('remark') or '').strip()

    if not (date_str and party_id and broker_id and amount):
        return JsonResponse({'error': 'Missing fields'}, status=400)

    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        amt = float(amount)
    except Exception:
        return JsonResponse({'error': 'Invalid input'}, status=400)

    # ⬇️ scoped lookups
    party = get_object_or_404(HeadParty, pk=party_id, org=request.current_org)
    broker = get_object_or_404(Broker, pk=broker_id, org=request.current_org)

    with transaction.atomic():
        daily_page, _ = DailyPage.objects.get_or_create(org=request.current_org, date=date_obj)
        entry = NaameEntry.objects.create(
            daily_page=daily_page,
            party=party,
            broker=broker,
            amount=amt,
            remark=remark
        )

    data = {
        'entry_no': entry.entry_no,
        'party_name': party.partyname,
        'broker_name': broker.brokername,
        'amount': f"{entry.amount:.2f}",
        'remark': entry.remark,
    }
    return JsonResponse({'success': True, 'entry': data})

@require_POST

def daily_page_jama_delete(request, entry_no):
    entry = get_object_or_404(JamaEntry, entry_no=entry_no, daily_page__org=request.current_org)
    entry.delete()
    return JsonResponse({'success': True, 'entry_no': entry_no})


@require_POST

def daily_page_naame_delete(request, entry_no):
    entry = get_object_or_404(NaameEntry, entry_no=entry_no, daily_page__org=request.current_org)
    entry.delete()
    return JsonResponse({'success': True, 'entry_no': entry_no})



def daily_page_pdf(request):
    date = request.GET.get('date')
    if not date:
        return HttpResponse("Date not provided", status=400)

    # Query entries
    jama_entries = list(JamaEntry.objects.filter(daily_page__date=date).order_by('entry_no'))
    naame_entries = list(NaameEntry.objects.filter(daily_page__date=date).order_by('entry_no'))

    total_jama = sum(float(j.amount or 0) for j in jama_entries)
    total_naame = sum(float(n.amount or 0) for n in naame_entries)
    diff = total_jama - total_naame

    # --- PDF setup (landscape A4) ---
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Daily Report - {date}", ln=True, align="C")
    pdf.ln(6)

    # Column layout (left and right table)
    # A4 landscape usable width ~= 277mm (after small margins). We'll split into two panels.
    left_x = 12
    right_x = 12 + 137 + 8  # left panel width + small gap
    panel_width = 137  # width for each table panel

    # Column widths inside a panel (sum <= panel_width)
    # no, party, broker, amount, remark
    w_no = 10
    w_party = 48
    w_broker = 38
    w_amount = 22
    w_remark = panel_width - (w_no + w_party + w_broker + w_amount)  # remaining

    # Header row height
    hdr_h = 8
    row_h = 8

    # Function to draw a single panel (list of entries)
    def draw_panel(x, y, title, entries):
        pdf.set_xy(x, y)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(panel_width, 6, title, ln=1)

        # table header
        pdf.set_font("Arial", "B", 10)
        pdf.set_xy(x, pdf.get_y())
        pdf.cell(w_no, hdr_h, "No", border=1, align="C")
        pdf.cell(w_party, hdr_h, "Party", border=1, align="L")
        pdf.cell(w_broker, hdr_h, "Broker", border=1, align="L")
        pdf.cell(w_amount, hdr_h, "Amount", border=1, align="R")
        pdf.cell(w_remark, hdr_h, "Remark", border=1, align="L")
        pdf.ln(hdr_h)

        # rows
        pdf.set_font("Arial", "", 10)
        for e in entries:
            # ensure we don't go beyond bottom margin - FPDF auto-adds page if needed
            pdf.set_x(x)
            pdf.cell(w_no, row_h, str(e.entry_no), border=1)
            # party (truncate if too long)
            party_text = getattr(e.party, "partyname", str(e.party) if e.party else "")
            if len(party_text) > 35:
                party_text = party_text[:32] + "..."
            pdf.cell(w_party, row_h, party_text, border=1)
            # broker
            broker_text = getattr(e.broker, "brokername", str(e.broker) if e.broker else "")
            if len(broker_text) > 30:
                broker_text = broker_text[:27] + "..."
            pdf.cell(w_broker, row_h, broker_text, border=1)
            # amount (right aligned)
            pdf.cell(w_amount, row_h, f"{float(e.amount or 0):.2f}", border=1, align="R")
            # remark - truncate
            remark_text = (e.remark or "")
            if len(remark_text) > 40:
                remark_text = remark_text[:37] + "..."
            pdf.cell(w_remark, row_h, remark_text, border=1)
            pdf.ln(row_h)

        # after rows: draw total under Amount column (aligned under amount cell)
        # Move to the footer row position (we'll draw a row showing 'Total' in the left columns and value under Amount)
        pdf.set_x(x)
        # create a cell spanning no+party+broker widths with label 'Total'
        span_width = w_no + w_party + w_broker
        pdf.set_font("Arial", "B", 10)
        pdf.cell(span_width, hdr_h, "Total", border='T')
        # amount cell with top border
        # compute the total for this panel
        panel_total = sum(float(ent.amount or 0) for ent in entries)
        pdf.cell(w_amount, hdr_h, f"{panel_total:.2f}", border='T', align="R")
        # empty remark cell
        pdf.cell(w_remark, hdr_h, "", border='T')
        pdf.ln(hdr_h + 4)  # small gap after table

    # Draw both panels side-by-side starting from current y
    start_y = pdf.get_y()
    draw_panel(left_x, start_y, "Jama", jama_entries)
    draw_panel(right_x, start_y, "Naame", naame_entries)

    # Summary line (below panels)
    pdf.set_font("Arial", "B", 12)
    # Put jama total on left area
    pdf.set_xy(left_x, pdf.get_y())
    pdf.cell(120, 8, f"Jama Total: {total_jama:.2f}", ln=0)
    # Put naame total on right area
    # place it roughly under right panel
    pdf.set_xy(right_x, pdf.get_y())
    pdf.cell(120, 8, f"Naame Total: {total_naame:.2f}", ln=0)
    pdf.ln(10)
    # Difference on the right aligned
    pdf.set_font("Arial", "B", 12)
    # align right towards page right margin
    page_right = pdf.w - 12
    diff_text = f"Difference (Jama - Naame): {diff:.2f}"
    text_width = pdf.get_string_width(diff_text) + 2
    pdf.set_xy(page_right - text_width, pdf.get_y())
    pdf.cell(text_width, 8, diff_text, ln=1, align='R')

    # Output and return
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="DailyReport_{date}.pdf"'
    return response




class AllPartyBalanceView(TemplateView):
    """
    Simplified AllPartyBalanceView matching compact template (no filters).
    Supports POST actions via buttons with name="action":
      - balance       : show table in page
      - print         : render printable HTML (user can browser-print)
      - export_excel  : return .xlsx (requires openpyxl)
      - pdf           : return PDF generated with fpdf2
    """
    template_name = "brokerapp/account/all_party_balance.html"
    printable_template = "brokerapp/account/all_party_balance_printable.html"

    # ---------- helpers ----------
    def _org_filter(self, qs):
        org_id = self.request.session.get("org_id")
        if not org_id:
            return qs
        field_names = [f.attname for f in qs.model._meta.fields]
        if "org_id" in field_names:
            return qs.filter(org_id=org_id)
        return qs

    def _sum(self, qs, field):
        """Safe sum returning Decimal(0) when None."""
        return qs.aggregate(t=Sum(field))["t"] or Decimal("0")

    # ---------- GET ----------
    def get(self, request, *args, **kwargs):
        today = date.today()
        ctx = self._build_context(start=today, end=today, party=None)
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- POST ----------
    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        today = date.today()

        # Build rows/totals (same data used by all actions)
        ctx = self._build_context(start=today, end=today, party=None)

        # Balance -> show table in same template
        if action == "balance" or not action:
            ctx["show_table"] = True
            return self.render_to_response(ctx)

        # Print -> render printable HTML (no buttons)
        if action == "print":
            ctx["show_table"] = True
            return render(request, self.printable_template, ctx)

        # Export Excel -> create .xlsx (requires openpyxl)
        if action == "export_excel":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                return HttpResponse(
                    "Required package 'openpyxl' not installed. Install with: pip install openpyxl",
                    content_type="text/plain",
                    status=500
                )

            wb = Workbook()
            ws = wb.active
            ws.title = "All Party Balance"

            headers = ["Party", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            ws.append(headers)

            for r in ctx["rows"]:
                pname = getattr(r["party"], "partyname", str(r["party"]))
                ws.append([
                    pname,
                    float(r["op_dr"]), float(r["op_cr"]),
                    float(r["opening"]), float(r["sale"]),
                    float(r["purchase"]), float(r["naame"]),
                    float(r["jama"]), float(r["balance"])
                ])

            # auto column width (simple)
            for i, col in enumerate(ws.columns, start=1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(
                out.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = f'attachment; filename="all_party_balance_{today}.xlsx"'
            return resp

        # PDF -> generate using fpdf2
        if action == "pdf":
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=10)

            # Header
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "All Party Balance", ln=True, align="C")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, f"Generated on: {today.strftime('%d-%m-%Y')}", ln=True, align="C")
            pdf.ln(4)

            # Table headers
            headers = ["Party", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            col_widths = [50, 18, 18, 24, 18, 22, 18, 18, 22]  # total should fit A4 width with margins

            pdf.set_font("Helvetica", "B", 9)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align="C")
            pdf.ln(8)

            # Rows
            pdf.set_font("Helvetica", "", 9)
            for r in ctx["rows"]:
                vals = [
                    getattr(r["party"], "partyname", str(r["party"])),
                    f"{r['op_dr']:.2f}", f"{r['op_cr']:.2f}",
                    f"{r['opening']:.2f}", f"{r['sale']:.2f}",
                    f"{r['purchase']:.2f}", f"{r['naame']:.2f}",
                    f"{r['jama']:.2f}", f"{r['balance']:.2f}"
                ]
                for i, v in enumerate(vals):
                    align = "L" if i == 0 else "R"
                    pdf.cell(col_widths[i], 7, v, border=1, align=align)
                pdf.ln(7)

            # Totals row
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_widths[0], 8, "TOTAL", border=1, align="L")
            totals = ctx["totals"]
            total_vals = [
                totals["opdr"], totals["opcr"], totals["sale"], totals["purchase"],
                totals["naame"], totals["jama"], totals["balance"]
            ]
            # place totals aligned under numeric columns (skip party col)
            # mapping to header indices: 1(OpDr),2(OpCr),3(Opening) etc. We'll print totals aligned with numeric columns.
            # For simplicity, print totals under Op Dr onward; keep Opening blank since it's derived per-party.
            pdf.cell(col_widths[1], 8, f"{totals['opdr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[2], 8, f"{totals['opcr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[3], 8, "", border=1, align="R")  # Opening total left blank
            pdf.cell(col_widths[4], 8, f"{totals['sale']:.2f}", border=1, align="R")
            pdf.cell(col_widths[5], 8, f"{totals['purchase']:.2f}", border=1, align="R")
            pdf.cell(col_widths[6], 8, f"{totals['naame']:.2f}", border=1, align="R")
            pdf.cell(col_widths[7], 8, f"{totals['jama']:.2f}", border=1, align="R")
            pdf.cell(col_widths[8], 8, f"{totals['balance']:.2f}", border=1, align="R")
            pdf.ln(10)

            buf = io.BytesIO()
            pdf.output(buf)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="all_party_balance_{today}.pdf"'
            return resp

        # Unknown action -> render without table
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- core calculation ----------
    def _build_context(self, start, end, party):
        # parties
        parties = HeadParty.objects.all().order_by("partyname")
        if party:
            parties = parties.filter(partyname=party.partyname)

        rows = []
        totals = {
            "opdr": Decimal("0"), "opcr": Decimal("0"),
            "sale": Decimal("0"), "purchase": Decimal("0"),
            "naame": Decimal("0"), "jama": Decimal("0"),
            "balance": Decimal("0")
        }

        dp_before = DailyPage.objects.filter(date__lt=start)
        dp_range = DailyPage.objects.filter(date__range=(start, end))

        org_id = self.request.session.get("org_id")
        if org_id:
            dp_before = dp_before.filter(org_id=org_id)
            dp_range = dp_range.filter(org_id=org_id)
            parties = parties.filter(org_id=org_id)

        for p in parties:
            op_dr = Decimal(getattr(p, "openingdebit", 0) or 0)
            op_cr = Decimal(getattr(p, "openingcredit", 0) or 0)

            sale_before = self._sum(self._org_filter(SaleMaster.objects.filter(party=p, invdate__lt=start)), "netamt")
            purch_before = self._sum(self._org_filter(PurchaseMaster.objects.filter(party=p, invdate__lt=start)), "netamt")
            naame_before = self._sum(NaameEntry.objects.filter(daily_page__in=dp_before, party=p), "amount")
            jama_before = self._sum(JamaEntry.objects.filter(daily_page__in=dp_before, party=p), "amount")

            opening = (op_dr - op_cr) + (sale_before - purch_before + naame_before - jama_before)

            sale = self._sum(self._org_filter(SaleMaster.objects.filter(party=p, invdate__range=(start, end))), "netamt")
            purchase = self._sum(self._org_filter(PurchaseMaster.objects.filter(party=p, invdate__range=(start, end))), "netamt")
            naame = self._sum(NaameEntry.objects.filter(daily_page__in=dp_range, party=p), "amount")
            jama = self._sum(JamaEntry.objects.filter(daily_page__in=dp_range, party=p), "amount")

            balance = opening + sale - purchase + naame - jama

            rows.append({
                "party": p, "op_dr": op_dr, "op_cr": op_cr, "opening": opening,
                "sale": sale, "purchase": purchase, "naame": naame,
                "jama": jama, "balance": balance
            })

            totals["opdr"] += op_dr
            totals["opcr"] += op_cr
            totals["sale"] += sale
            totals["purchase"] += purchase
            totals["naame"] += naame
            totals["jama"] += jama
            totals["balance"] += balance

        return {"rows": rows, "totals": totals, "start": start, "end": end}

def party_statement(request):
    """
    Party Statement — no date filters.
    Select a party from dropdown and click 'Statement' to load its entries.
    Uses:
     - HeadParty.openingdebit / openingcredit
     - SaleMaster.netamt (treated as Debit)
     - PurchaseMaster.netamt (treated as Credit)
     - NaameEntry.amount (Debit)
     - JamaEntry.amount (Credit)
    """
    party_id = request.GET.get('party')  # using HeadParty.partyname as PK
    parties = HeadParty.objects.order_by('partyname')
    entries = []

    if party_id:
        head = get_object_or_404(HeadParty, pk=party_id)

        # Opening
        if head.openingdebit and head.openingdebit != Decimal('0'):
            entries.append({
                'entry_no': 'OPEN',
                'date': None,
                'debit': head.openingdebit,
                'credit': Decimal('0'),
                'remark': 'Opening (Dr)'
            })
        elif head.openingcredit and head.openingcredit != Decimal('0'):
            entries.append({
                'entry_no': 'OPEN',
                'date': None,
                'debit': Decimal('0'),
                'credit': head.openingcredit,
                'remark': 'Opening (Cr)'
            })

        # Sales -> Debit
        for s in SaleMaster.objects.filter(party=head).order_by('invdate'):
            entries.append({
                'entry_no': s.invno,
                'date': s.invdate,
                'debit': s.netamt,
                'credit': Decimal('0'),
                'remark': s.remark or f"Sale Inv#{s.invno}"
            })

        # Purchases -> Credit
        for p in PurchaseMaster.objects.filter(party=head).order_by('invdate'):
            entries.append({
                'entry_no': p.invno,
                'date': p.invdate,
                'debit': Decimal('0'),
                'credit': p.netamt,
                'remark': p.remark or f"Purchase Inv#{p.invno}"
            })

        # Naame entries -> Debit
        for n in NaameEntry.objects.filter(party=head).order_by('daily_page__date'):
            entries.append({
                'entry_no': n.entry_no,
                'date': n.daily_page.date,
                'debit': n.amount,
                'credit': Decimal('0'),
                'remark': n.remark or 'Naame'
            })

        # Jama entries -> Credit
        for j in JamaEntry.objects.filter(party=head).order_by('daily_page__date'):
            entries.append({
                'entry_no': j.entry_no,
                'date': j.daily_page.date,
                'debit': Decimal('0'),
                'credit': j.amount,
                'remark': j.remark or 'Jama'
            })

        # Sort entries by date (opening has date None)
        entries = sorted(entries, key=lambda x: (x['date'] is None, x['date'] or ''))

        # totals + running balance
        total_debit = sum(e['debit'] for e in entries)
        total_credit = sum(e['credit'] for e in entries)
        bal = Decimal('0')
        for e in entries:
            bal += (e['debit'] or Decimal('0')) - (e['credit'] or Decimal('0'))
            e['balance'] = bal

        balance = total_debit - total_credit

    else:
        head = None
        total_debit = total_credit = balance = Decimal('0')

    return render(request, 'brokerapp/account/party_statement.html', {
        'parties': parties,
        'selected': head,
        'entries': entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'balance': balance,
    })
    
def broker_statement(request):
    """
    Broker statement built from multiple models that reference Broker:
      - JamaEntry  => treated as CREDIT (money received)
      - NaameEntry => treated as DEBIT  (money paid)
      - SaleMaster => uses 'dramt' as DEBIT by default (adjust if needed)
      - PurchaseMaster => uses 'dramt' as CREDIT by default (adjust if needed)

    If any mapping is different for your business logic, update the mapping lines below.
    """
    brokers = Broker.objects.all().order_by('brokername')
    selected = None
    entries = []
    total_debit = total_credit = 0.0
    balance = 0.0

    broker_key = request.GET.get('broker')  # brokername is primary_key on Broker
    if broker_key:
        selected = get_object_or_404(Broker, brokername=broker_key)

        # 1) JamaEntry -> treat as CREDIT (amount credited to account)
        jama_qs = JamaEntry.objects.filter(broker=selected).order_by('created_at')
        for j in jama_qs:
            date_val = j.created_at.date() if getattr(j, 'created_at', None) else None
            amount = float(j.amount or 0)
            entries.append({
                'entry_no': f"J-{j.entry_no}",
                'date': date_val,
                'debit': 0.0,
                'credit': amount,
                'remark': (j.remark or '') + f" (Jama)",
            })

        # 2) NaameEntry -> treat as DEBIT (amount paid)
        naame_qs = NaameEntry.objects.filter(broker=selected).order_by('created_at')
        for n in naame_qs:
            date_val = n.created_at.date() if getattr(n, 'created_at', None) else None
            amount = float(n.amount or 0)
            entries.append({
                'entry_no': f"N-{n.entry_no}",
                'date': date_val,
                'debit': amount,
                'credit': 0.0,
                'remark': (n.remark or '') + f" (Naame)",
            })

        # 3) SaleMaster -> default: use dramt as DEBIT (broker charge). Change if business logic differs.
        sale_qs = SaleMaster.objects.filter(broker=selected).order_by('invdate')
        for s in sale_qs:
            date_val = getattr(s, 'invdate', None)
            # prefer dramt (dalali amount). If you want different field, change below.
            amt = float(getattr(s, 'dramt', 0) or 0)
            # If your SaleMaster stores dalali differently (e.g. 'dr' or 'dramt'), adjust here.
            entries.append({
                'entry_no': f"S-{getattr(s, 'invno', '')}",
                'date': date_val,
                'debit': amt,
                'credit': 0.0,
                'remark': (getattr(s, 'remark', '') or '') + f" (Sale)",
            })

        # 4) PurchaseMaster -> default: use dramt as CREDIT (adjust if needed)
        purchase_qs = PurchaseMaster.objects.filter(broker=selected).order_by('invdate')
        for p in purchase_qs:
            date_val = getattr(p, 'invdate', None)
            amt = float(getattr(p, 'dramt', 0) or 0)
            entries.append({
                'entry_no': f"P-{getattr(p, 'invno', '')}",
                'date': date_val,
                'debit': 0.0,
                'credit': amt,
                'remark': (getattr(p, 'remark', '') or '') + f" (Purchase)",
            })

        # Sort combined entries by date (None last) then by entry_no
        # Convert None dates to far future for stable sort
        def _date_key(item):
            d = item.get('date')
            return (d or datetime.max.date(), item.get('entry_no'))

        entries.sort(key=_date_key)

        # Compute totals and running balance (balance = previous + debit - credit)
        total_debit = sum(item['debit'] for item in entries)
        total_credit = sum(item['credit'] for item in entries)
        running = 0.0
        for item in entries:
            running += (item['debit'] - item['credit'])
            # attach running balance to item for template
            item['balance'] = running

        balance = running

    context = {
        'brokers': brokers,
        'selected': selected,
        'entries': entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'balance': balance,
    }
    return render(request, 'brokerapp/account/broker_statement.html', context)
    